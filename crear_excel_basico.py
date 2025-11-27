# crear_excel_basico.py
"""
Script para crear un archivo Excel básico para el dashboard
Solo para ayudar a los estudiantes a empezar
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_FILE

def crear_excel_template():
    """Crear plantilla básica de Excel"""
    print("📊 Creando plantilla de Excel...")
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Crear hojas necesarias
    hojas = [
        "Dashboard",
        "Datos Diarios", 
        "Top Peliculas",
        "Por Categoria"
    ]
    
    for nombre_hoja in hojas:
        ws = wb.create_sheet(nombre_hoja)
        
        # Agregar encabezado bonito
        ws['A1'] = f"Hoja: {nombre_hoja}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:D1')
        
        if nombre_hoja == "Dashboard":
            ws['A3'] = "INSTRUCCIONES:"
            ws['A3'].font = Font(bold=True)
            ws['A4'] = "1. Ejecuta: python sakila_automation.py"
            ws['A5'] = "2. Importa los CSV desde carpeta 'output/'"
            ws['A6'] = "3. Crea tablas dinámicas y gráficos aquí"
            ws['A7'] = "4. Para actualizar: Alt + F5 en cualquier tabla dinámica"
        
        elif nombre_hoja == "Datos Diarios":
            ws['A3'] = "Importar archivo: output/resumen_diario.csv"
            ws['A3'].font = Font(italic=True, color="808080")
            ws['A5'] = "Fecha"
            ws['B5'] = "Ingresos"
            ws['C5'] = "Clientes"
            ws['D5'] = "Rentas"
            for col in ['A5', 'B5', 'C5', 'D5']:
                ws[col].font = Font(bold=True)
                ws[col].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        elif nombre_hoja == "Top Peliculas":
            ws['A3'] = "Importar archivo: output/top_peliculas.csv"
            ws['A3'].font = Font(italic=True, color="808080")
            ws['A5'] = "Pelicula"
            ws['B5'] = "Ingresos"
            ws['C5'] = "Total_Rentas"
            for col in ['A5', 'B5', 'C5']:
                ws[col].font = Font(bold=True)
                ws[col].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        elif nombre_hoja == "Por Categoria":
            ws['A3'] = "Importar archivo: output/por_categoria.csv"
            ws['A3'].font = Font(italic=True, color="808080")
            ws['A5'] = "Categoria"
            ws['B5'] = "Ingresos"
            ws['C5'] = "Total_Rentas"
            for col in ['A5', 'B5', 'C5']:
                ws[col].font = Font(bold=True)
                ws[col].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    # Guardar archivo
    wb.save(EXCEL_FILE)
    print(f"✓ Archivo Excel creado: {EXCEL_FILE}")
    print("\n📋 Hojas creadas:")
    for hoja in hojas:
        print(f"   • {hoja}")
    
    print("\n💡 Próximos pasos:")
    print("   1. Ejecuta: python automatizacion_simple.py")
    print("   2. Abre el Excel y sigue las instrucciones de cada hoja")
    print("   3. Importa los CSV y crea tus gráficos")

if __name__ == "__main__":
    try:
        crear_excel_template()
    except Exception as e:
        print(f"❌ Error: {e}")
        print("\nAsegúrate de tener instalado: pip install openpyxl")

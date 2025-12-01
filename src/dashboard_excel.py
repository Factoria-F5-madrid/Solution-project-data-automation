# src/dashboard_excel.py
# Generador automático de dashboards en Excel con tablas dinámicas
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from .config import *

class DashboardExcel:
    def __init__(self, output_file="dashboard_sakila.xlsx"):
        self.output_file = f"{OUTPUT_FOLDER}/{output_file}"
        self.workbook = openpyxl.Workbook()
        
        # Intentar eliminar archivo existente si existe
        self._eliminar_archivo_existente()
        
    def _eliminar_archivo_existente(self):
        """Eliminar archivo Excel existente si existe"""
        try:
            if os.path.exists(self.output_file):
                os.remove(self.output_file)
                print(f"✓ Archivo existente eliminado: {self.output_file}")
        except PermissionError:
            print(f"⚠️  ADVERTENCIA: No se puede eliminar {self.output_file}")
            print("   Probablemente está abierto en Excel. Ciérralo y vuelve a intentar.")
            # Intentar con nombre alternativo
            import time
            timestamp = int(time.time())
            self.output_file = f"{OUTPUT_FOLDER}/dashboard_sakila_{timestamp}.xlsx"
            print(f"   Creando archivo alternativo: {self.output_file}")
        except Exception as e:
            print(f"⚠️  Error eliminando archivo existente: {e}")
            # Continuar con nombre alternativo
            import time
            timestamp = int(time.time())
            self.output_file = f"{OUTPUT_FOLDER}/dashboard_sakila_{timestamp}.xlsx"
        
    def verificar_archivo_disponible(self):
        """Verificar si el archivo está disponible para escritura"""
        if not os.path.exists(self.output_file):
            return True
            
        try:
            # Intentar abrir el archivo en modo escritura
            with open(self.output_file, 'r+b') as f:
                pass
            return True
        except PermissionError:
            return False
        except Exception:
            return True  # Si es otro error, intentar continuar
    
    def cargar_datos_principales(self):
        """Cargar todos los datasets generados por el ETL"""
        print("📊 Cargando datos desde ETL...")
        
        # Ejecutar ETL y obtener todos los datasets
        from .sakila_etl import proceso_completo
        datasets = proceso_completo()
        
        if datasets:
            self.datos = datasets['datos_principales']
            self.resumen_fechas = datasets['resumen_fechas']
            self.por_pais = datasets['por_pais']
            self.por_ciudad = datasets['por_ciudad']
            self.clientes_resumen = datasets['clientes_resumen']
            self.por_mes = datasets['por_mes']
            
            print(f"✓ {len(self.datos)} registros cargados con 5 datasets adicionales")
            return True
        else:
            print("❌ Error: No se pudieron cargar los datos del ETL")
            return False
    
    def crear_hoja_con_datos(self, nombre_hoja, dataframe, descripcion):
        """Crear hoja formateada con cualquier dataset"""
        ws = self.workbook.create_sheet(nombre_hoja)
        
        # Título descriptivo
        ws['A1'] = descripcion
        ws['A1'].font = Font(size=12, bold=True, color="366092")
        ws.merge_cells(f'A1:{get_column_letter(len(dataframe.columns))}1')
        
        # Agregar datos desde la fila 3
        for r_idx, r in enumerate(dataframe_to_rows(dataframe, index=False, header=True), start=3):
            for c_idx, value in enumerate(r, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Formatear encabezados (fila 3)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(dataframe.columns) + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Formatear como tabla de Excel
        tabla_rango = f"A3:{get_column_letter(len(dataframe.columns))}{len(dataframe) + 3}"
        tabla = Table(displayName=f"Tabla{nombre_hoja.replace('_', '')}", ref=tabla_rango)
        
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=True
        )
        tabla.tableStyleInfo = style
        ws.add_table(tabla)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def crear_dashboard_resumen(self):
        """Crear hoja de dashboard principal con métricas clave"""
        ws = self.workbook.create_sheet("Dashboard")
        
        # Título principal
        ws['A1'] = "DASHBOARD SAKILA - MÉTRICAS PRINCIPALES"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Métricas principales calculadas
        total_ingresos = self.datos['monto'].sum()
        total_clientes = self.datos['customer_id'].nunique()
        total_transacciones = len(self.datos)
        paises_unicos = self.datos['country'].nunique()
        ciudades_unicas = self.datos['city'].nunique()
        
        # Crear métricas principales
        metricas = [
            ["📊 MÉTRICAS GENERALES", ""],
            ["Total Ingresos", f"${total_ingresos:,.2f}"],
            ["Clientes Únicos", f"{total_clientes:,}"],
            ["Total Transacciones", f"{total_transacciones:,}"],
            ["Países Activos", f"{paises_unicos}"],
            ["Ciudades Activas", f"{ciudades_unicas}"],
            ["Ingreso/Transacción", f"${total_ingresos/total_transacciones:,.2f}"]
        ]
        
        # Agregar métricas con formato
        for i, metrica in enumerate(metricas, start=3):
            ws[f'A{i}'] = metrica[0]
            ws[f'B{i}'] = metrica[1]
            
            if i == 3:  # Encabezado
                ws[f'A{i}'].font = Font(bold=True, size=12, color="FFFFFF")
                ws[f'B{i}'].font = Font(bold=True, size=12, color="FFFFFF")
                ws[f'A{i}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws[f'B{i}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            else:
                ws[f'A{i}'].font = Font(bold=True)
        
        # Top 5 países resumido
        top_paises = self.datos.groupby('country')['monto'].sum().sort_values(ascending=False).head(5)
        
        ws['D3'] = "� TOP 5 PAÍSES"
        ws['D3'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['D3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('D3:E3')
        
        for i, (pais, ingresos) in enumerate(top_paises.items(), start=4):
            ws[f'D{i}'] = pais
            ws[f'E{i}'] = f"${ingresos:,.0f}"
            ws[f'D{i}'].font = Font(bold=True)
        
        # Instrucciones
        ws['A11'] = "� NOTAS PARA EL DASHBOARD:"
        ws['A11'].font = Font(bold=True, size=12, color="FF6600")
        
        notas = [
            "• Cada hoja contiene datos listos para gráficos específicos",
            "• Usa 'Resumen_Fechas' para gráficos temporales (líneas)",
            "• Usa 'Por_Pais' y 'Por_Ciudad' para gráficos de barras",
            "• 'Top_Clientes' para análisis de clientes VIP",
            "• 'Datos_Principales' para Power Query y análisis personalizado"
        ]
        
        for i, nota in enumerate(notas, start=12):
            ws[f'A{i}'] = nota
            ws[f'A{i}'].font = Font(size=10)
        
        return ws
    
    def generar_dashboard_completo(self):
        """Generar Excel con múltiples hojas de datos pre-procesados"""
        print("🔄 Generando Excel con datos pre-procesados...")
        
        # Verificar disponibilidad del archivo
        if not self.verificar_archivo_disponible():
            print("⚠️  ADVERTENCIA: El archivo Excel está abierto en otra aplicación.")
            print("   Cierra Excel y vuelve a intentar, o se creará con nombre alternativo.")
        
        if not self.cargar_datos_principales():
            return False
        
        try:
            # Eliminar hoja por defecto
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # 1. Crear hoja de datos principales (índice 0)
            self.crear_hoja_con_datos(
                "Datos_Principales", 
                self.datos, 
                "DATOS PRINCIPALES - Fuente completa para análisis y Power Query"
            )
            
            # 2. Crear hojas con datos pre-procesados (listas para gráficos)
            self.crear_hoja_con_datos(
                "Resumen_Fechas", 
                self.resumen_fechas, 
                "RESUMEN POR FECHAS - Datos diarios listos para gráfico de líneas temporal"
            )
            
            self.crear_hoja_con_datos(
                "Por_Pais", 
                self.por_pais, 
                "ANÁLISIS POR PAÍS - Datos agregados listos para gráfico de barras y mapas"
            )
            
            self.crear_hoja_con_datos(
                "Por_Ciudad", 
                self.por_ciudad, 
                "TOP CIUDADES - Top 20 ciudades listas para gráfico de barras"
            )
            
            self.crear_hoja_con_datos(
                "Top_Clientes", 
                self.clientes_resumen.head(50), 
                "TOP 50 CLIENTES - Clientes con mayor gasto para análisis detallado"
            )
            
            self.crear_hoja_con_datos(
                "Resumen_Mensual", 
                self.por_mes, 
                "RESUMEN MENSUAL - Tendencias por mes/año para análisis temporal"
            )
            
            # 3. Crear dashboard resumen con métricas (última hoja)
            self.crear_dashboard_resumen()
            
            # Guardar archivo
            try:
                self.workbook.save(self.output_file)
                print(f"✓ Excel creado exitosamente: {self.output_file}")
            except PermissionError:
                # Si sigue dando error, crear con timestamp
                import time
                timestamp = int(time.time())
                nuevo_nombre = f"{OUTPUT_FOLDER}/dashboard_sakila_{timestamp}.xlsx"
                self.workbook.save(nuevo_nombre)
                print(f"✓ Excel creado con nombre alternativo: {nuevo_nombre}")
                print("⚠️  El archivo original estaba en uso. Cierra Excel para la próxima ejecución.")
            
            print("✅ HOJAS CREADAS:")
            print("  📊 Datos_Principales - Fuente completa")
            print("  📈 Resumen_Fechas - Para gráfico temporal")
            print("  🌍 Por_Pais - Para gráfico de países")
            print("  🏙️ Por_Ciudad - Para top ciudades")
            print("  👥 Top_Clientes - Para análisis de clientes")
            print("  📅 Resumen_Mensual - Para tendencias")
            print("  📋 Dashboard - Métricas principales")
            print("\n🎯 RECOMENDACIÓN:")
            print("  1. Diseña tu dashboard manualmente en Excel usando estas hojas")
            print("  2. Para actualizar: ejecuta 'python main.py' y se actualizarán todos los datos")
            print("  3. Tu diseño se mantendrá, solo se actualizarán los datos")
            return True
            
        except Exception as e:
            print(f"❌ Error creando Excel: {e}")
            return False

def crear_dashboard():
    """Función principal para crear el dashboard"""
    dashboard = DashboardExcel()
    return dashboard.generar_dashboard_completo()

if __name__ == "__main__":
    crear_dashboard()